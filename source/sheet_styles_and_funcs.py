from openpyxl.cell import MergedCell
from openpyxl.utils import get_column_letter
from openpyxl.styles import (
    Font,
    Side,
    Border,
    Alignment
)

from source.credentials import styles


# Defines the width, font styles, alignments and border style dictionaries
# for the cells of table, taken from the JSON file 'styles_config'
dims = styles["dims"]
font = styles["font"]
alignment = styles["alignment"]
borders = styles["border_style"]


# Custom function which defines the border style for the current cell
def make_border(style) -> Border:
    cfg = borders[style]
    return Border(
        top=Side(border_style=cfg["top"] if "top" in cfg else None),
        left=Side(border_style=cfg["left"] if "left" in cfg else None),
        right=Side(border_style=cfg["right"] if "right" in cfg else None),
        bottom=Side(border_style=cfg["bottom"] if "bottom" in cfg else None),
    )

# Defines the font style for some cells in table
font_style = Font(**font["default"])

# Defines the bold font style for some cells in table
font_style_bold = Font(**font["bold_font"])

# Defines the default border for cell, where the top,
# left, right and bottom sides filled with medium thickness
border = make_border("default")

# Defines the styles for border of the cell
# where the bottom border must not be thick
no_bottom_border = make_border("no_bottom")

# Defines the styles for border of the cell
# where the top border must not be thick
no_top_border = make_border("no_top")

# Defines the styles for border of the cell
# where the top and bottom borders must not be thick
no_tb_border = make_border("no_top_bottom")

# Defines the center horizontal and vertical alignment
# inside the cell
mid_align = Alignment(**alignment["center"])

# Defines the left horizontal and center vertical alignment
# inside the cell
left_align = Alignment(**alignment["left"])


# Making a list of the tuples with the 'start' and 'end' coordinates
# of the grouped department
def merge_intervals(start: int, end: int, ws) -> list:
    edges = []
    for row in range(start + 1, end + 1):
        if ws[f"A{row}"].value != ws[f"A{start}"].value:
            edges.append((start, row - 1))
            start = row

    edges.append((start, end))

    return edges


# Main function which accepts the written above styles for the table page
def main_styles_acceptation(edges: list, ws) -> None:
    for start, end in edges:
        ws.merge_cells(f"A{start}:A{end}")
        ws.merge_cells(f"B{start}:B{end}")
        ws.row_dimensions[2].height = 20

        for row in ws.iter_rows(
                min_row=1,
                max_row=ws.max_row,
                min_col=1,
                max_col=ws.max_column
        ):
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue

                cell.font = font_style
                cell.alignment = mid_align

        for row in ws.iter_rows(
            min_row=1,
            max_row=1,
            min_col=1,
            max_col=ws.max_column
        ):
            for cell in row:
                cell.font = font_style_bold
        for row in ws.iter_rows(
            min_row=2,
            max_row=ws.max_row,
            min_col=ws.max_column,
            max_col=ws.max_column
        ):
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue
                cell.alignment = left_align


# The function chooses and accepts the border styles for the current cell
def border_styles(edges: list, ws) -> None:
    for start, end in edges:
        for column in range(ws.min_column, ws.max_column + 1):
            column_letter = get_column_letter(column)
            if isinstance(column_letter, MergedCell):
                continue

            ws[f"{column_letter}{start}"].border = no_bottom_border
            ws[f"{column_letter}{end}"].border = no_top_border
            for number in range(start + 1, end):
                ws[f"{column_letter}{number}"].border = no_tb_border


# Adding the last cells in the columns 'A' and 'B' which are
# consist the additional information of the table
def add_last_stats(end: int, ws) -> None:
    ws[f"A{end + 1}"] = "Всього: "
    ws[f"B{end + 1}"] = sum(
        row.value for row in ws["B"][1:] if row.value is not None
    )
    ws[f"A{ws.max_row}"].font = font_style_bold
    ws[f"B{ws.max_row}"].font = font_style_bold
    ws[f"A{ws.max_row}"].border = border
    ws[f"B{ws.max_row}"].alignment = mid_align
    ws[f"B{ws.max_row}"].border = border
