from datetime import date
from openpyxl import load_workbook

from source.report_functions import report_merge
from source.sheet_styles_and_funcs import (
    dims,
    merge_intervals,
    main_styles_acceptation,
    border_styles, add_last_stats
)


# The main function which executes editing the raw data,
# making the file and accepting the styles for the page
# of the final version file
def groups_edges() -> None:
    report_merge()
    report = f"./reports/Безпечне місто {date.today()}.xlsx"
    wb = load_workbook(report)
    ws = wb["Користувачі"]

    start = 2
    end = ws.max_row

    edges = merge_intervals(start, end, ws)
    for letter, width in dims.items():
        ws.column_dimensions[letter].width = width

    main_styles_acceptation(edges, ws)
    border_styles(edges, ws)

    add_last_stats(end, ws)

    wb.save(report)
