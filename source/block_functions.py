import os
from datetime import date, datetime
import pandas as pd
from openpyxl import load_workbook
from sqlalchemy import update, Table, MetaData
from sqlalchemy.exc import SQLAlchemyError

from source.report_functions import department_replacements
from source.credentials import (
    engine,
    request,
    styles
)
from source.sheet_styles_and_funcs import (
    merge_intervals,
    main_styles_acceptation,
    border_styles,
)


# Saving the raw data to Excel file
def save_block_to_excel(report) -> None:
    os.makedirs("blocked", exist_ok=True)
    report.to_excel(
        f"./blocked/Blocked inactive users {date.today()}.xlsx",
        index=False,
        sheet_name="Users"
    )


# Getting data about inactive users and changes block status
# for these users in database
def update_and_block(users_report: pd.DataFrame) -> None:
    meta = MetaData()
    users_table = Table(
        "cam_users",
        meta,
        autoload_with=engine
    )
    block_usernames = users_report["username"].to_list()
    if not block_usernames:
        return

    statement = (
        update(users_table)
        .where(users_table.c.username.in_(block_usernames))
        .values(block=1)
    )

    try:
        with engine.begin() as conn:
            conn.execute(statement)
    except SQLAlchemyError as error:
        raise error


# Gathering, merging and saving raw data about inactive users to Excel file
def block_report_merge_and_save() -> None:
    report = pd.read_sql(request, con=engine)
    report = report.rename(
        columns={
            "title": "department",
            "block": "block_status",
            "lastvisitDate": "last_visit",
        }
    )
    department_replacements(report)
    report["block_status"] = report["block_status"].replace(0, 1)
    report.insert(
        report.columns.get_loc("block_status") + 1,
        "block_date",
        datetime.today()
    )
    final_report = report [
        [
            "department",
            "username",
            "name",
            "block_status",
            "last_visit",
            "block_date"
        ]
    ]

    update_and_block(final_report.copy())

    save_block_to_excel(final_report)


# Main function, which edit saved file with Excel-styles.
def block_inactive() -> None:
    block_report_merge_and_save()
    report = f"./blocked/Blocked inactive users {date.today()}.xlsx"
    wb = load_workbook(report)
    ws = wb["Users"]

    start = 2
    end = ws.max_row

    edges = merge_intervals(start, end, ws)
    block_dims = styles["block_dims"]
    for letter, width in block_dims.items():
        ws.column_dimensions[letter].width = width

    main_styles_acceptation(edges, ws)
    border_styles(edges, ws)

    wb.save(report)
