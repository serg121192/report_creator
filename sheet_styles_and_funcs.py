from openpyxl.cell import MergedCell
from openpyxl.utils import get_column_letter
from openpyxl.styles import (
    Font,
    Side,
    Border,
    Alignment
)


# Defines the width for the cells in the columns of the table
dims = {
    "A" : 30,
    "B" : 10,
    "C" : 20,
    "D" : 15,
    "E" : 45,
}

# Defines the font style for some cells in table
font_style = Font(
    name="Times New Roman",
    size=11,
    bold=False,
)

# Defines the bold font style for some cells in table
font_style_bold = Font(
    name="Times New Roman",
    size=11,
    bold=True,
)

# Defines the thick style for borderline
thick = Side(
    border_style="medium",
    color="FF000000"
)

# Defines the styles for borders of the cell
border = Border(
    top=thick,
    left=thick,
    right=thick,
    bottom=thick,
)

# Defines the styles for border of the cell
# where the bottom border must not be thick
no_bottom_border = Border(
    top=thick,
    left=thick,
    right=thick,
)

# Defines the styles for border of the cell
# where the top border must not be thick
no_top_border = Border(
    left=thick,
    right=thick,
    bottom=thick,
)

# Defines the styles for border of the cell
# where the top and bottom borders must not be thick
no_tb_border = Border(
    left=thick,
    right=thick,
)

# Defines the center horizontal and vertical alignment
# inside the cell
mid_align = Alignment(
    vertical="center",
    horizontal="center",
    wrap_text=True,
)

# Defines the left horizontal and center vertical alignment
# inside the cell
left_align = Alignment(
    vertical="center",
    horizontal="left",
    wrap_text=True,
)


# Making a list of the tuples with the start and end coordinates
# of the grouped department
def merge_intervals(start: int, end: int, ws) -> list:
    _edges = []
    for row in range(start + 1, end + 1):
        if ws[f"A{row}"].value != ws[f"A{start}"].value:
            _edges.append((start, row - 1))
            start = row

    _edges.append((start, end))

    return _edges


# Main function which accepts the written above styles for the table page
def main_styles_acceptation(_edges: list, ws) -> None:
    for start, end in _edges:
        ws.merge_cells(f"A{start}:A{end}")
        ws.merge_cells(f"B{start}:B{end}")
        ws.row_dimensions[2].height = 20

        for row in ws.iter_rows(
                min_row=1,
                max_row=ws.max_row,
                min_col=1,
                max_col=ws.max_column
        ):
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue

                cell.font = font_style
                cell.alignment = mid_align

        for row in ws.iter_rows(
            min_row=1,
            max_row=1,
            min_col=1,
            max_col=ws.max_column
        ):
            for cell in row:
                cell.font = font_style_bold
        for row in ws.iter_rows(
            min_row=2,
            max_row=ws.max_row,
            min_col=ws.max_column,
            max_col=ws.max_column
        ):
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue
                cell.alignment = left_align


# Defines and accepts the border styles for the cells
def border_styles(edges: list, ws) -> None:
    for start, end in edges:
        for column in range(ws.min_column, ws.max_column + 1):
            column_letter = get_column_letter(column)
            if isinstance(column_letter, MergedCell):
                continue

            ws[f"{column_letter}{start}"].border = no_bottom_border
            ws[f"{column_letter}{end}"].border = no_top_border
            for number in range(start + 1, end):
                ws[f"{column_letter}{number}"].border = no_tb_border


# Adding the last cells in the columns 'A' and 'B' which are
# consist the complex information of the table
def add_last_stats(end: int, ws) -> None:
    ws[f"A{end + 1}"] = "Всього: "
    ws[f"B{end + 1}"] = sum(
        row.value for row in ws["B"][1:] if row.value is not None
    )
    ws[f"A{ws.max_row}"].font = font_style_bold
    ws[f"B{ws.max_row}"].font = font_style_bold
    ws[f"A{ws.max_row}"].border = border
    ws[f"B{ws.max_row}"].alignment = mid_align
    ws[f"B{ws.max_row}"].border = border
