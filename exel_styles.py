from datetime import date
from openpyxl import load_workbook

from report_functions import report_merge
from sheet_styles_and_funcs import (
    dims,
    merge_intervals,
    main_styles_acceptation,
    border_styles, add_last_stats
)


def _groups_edges():
    report_merge()
    report = f"Безпечне місто {date.today()}.xlsx"
    wb = load_workbook(report)
    ws = wb["Користувачі"]

    start = 2
    end = ws.max_row

    _edges = merge_intervals(start, end, ws)
    for letter, width in dims.items():
        ws.column_dimensions[letter].width = width

    main_styles_acceptation(_edges, ws)
    border_styles(_edges, ws)

    add_last_stats(end, ws)

    wb.save(report)
